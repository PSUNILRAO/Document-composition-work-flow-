"""
data_loader.py
──────────────
Reads data from Excel, CSV, or database.
Validates every row against schema.json field contracts.
Returns clean, typed Python dicts ready for the renderer.
"""

import json
import csv
import re
import logging
from pathlib import Path
from datetime import datetime, date
from typing import Iterator

import openpyxl

log = logging.getLogger(__name__)

CONFIG_DIR  = Path(__file__).parent / "config"
SCHEMA_PATH = CONFIG_DIR / "schema.json"
# Backwards-compatibility: schema.json used to live at the repo root; fall back
# to that location when the config/ directory is not present.
if not SCHEMA_PATH.exists():
    _root_schema = Path(__file__).parent / "schema.json"
    if _root_schema.exists():
        SCHEMA_PATH = _root_schema


# ── Schema loader ─────────────────────────────────────────────────────────────
_schema_cache: dict | None = None

def load_schema() -> dict:
    global _schema_cache
    if _schema_cache is None:
        with open(SCHEMA_PATH, encoding="utf-8") as f:
            _schema_cache = json.load(f)
    return _schema_cache

def get_doc_schema(doc_type: str) -> dict:
    schema = load_schema()
    if doc_type not in schema["document_types"]:
        raise ValueError(f"Unknown doc_type '{doc_type}'. "
                         f"Available: {list(schema['document_types'].keys())}")
    return schema["document_types"][doc_type]


# ── Type coercion ─────────────────────────────────────────────────────────────
def _to_float(v) -> float:
    if v is None or v == "":
        return 0.0
    return float(str(v).replace(",", "").replace("$", "").strip())

def _to_date(v) -> str:
    """Return date as string; normalise various input formats."""
    if v is None or v == "":
        return ""
    if isinstance(v, (datetime, date)):
        return v.strftime("%d %b %Y")
    s = str(v).strip()
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%m/%d/%Y", "%d-%m-%Y",
                "%d %b %Y", "%B %d, %Y", "%d %B %Y"):
        try:
            return datetime.strptime(s, fmt).strftime("%d %b %Y")
        except ValueError:
            pass
    return s  # return as-is if unparseable

def _coerce(value, field_def: dict):
    ftype = field_def.get("type", "string")
    if ftype == "currency":
        return _to_float(value)
    if ftype == "number":
        return _to_float(value)
    if ftype == "percent":
        v = _to_float(value)
        return v / 100 if v > 1 else v   # normalise 5.5 → 0.055
    if ftype == "date":
        return _to_date(value)
    if ftype == "boolean":
        return str(value).strip().lower() in ("true", "yes", "1", "y")
    return str(value).strip() if value is not None else ""


# ── Row validation ────────────────────────────────────────────────────────────
class ValidationError(Exception):
    def __init__(self, errors: list[str]):
        self.errors = errors
        super().__init__("\n".join(errors))

def validate_row(row: dict, fields: dict, row_num: int = 0) -> list[str]:
    errors = []
    for field_name, field_def in fields.items():
        if field_def.get("type") == "list":
            continue
        value = row.get(field_name)
        if field_def.get("required") and (value is None or str(value).strip() == ""):
            errors.append(f"Row {row_num}: Required field '{field_name}' is missing or empty.")
            continue
        allowed = field_def.get("allowed")
        if allowed and value and str(value).strip() not in allowed:
            errors.append(f"Row {row_num}: Field '{field_name}' value '{value}' "
                          f"not in allowed list {allowed}.")
    return errors


# ── Data masking ──────────────────────────────────────────────────────────────
def mask_field(value: str, mask_pattern: str) -> str:
    """Apply masking pattern like '****{{last4}}' or '{{first5}}****{{last1}}'."""
    if not value:
        return value
    v = str(value)
    result = mask_pattern
    result = re.sub(r"\{\{last(\d+)\}\}",
                    lambda m: v[-int(m.group(1)):], result)
    result = re.sub(r"\{\{first(\d+)\}\}",
                    lambda m: v[:int(m.group(1))], result)
    return result


# ── Excel reader ──────────────────────────────────────────────────────────────
def _read_excel_sheet(path: str, sheet_name: str | None = None) -> list[dict]:
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    ws = wb[sheet_name] if sheet_name and sheet_name in wb.sheetnames else wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    headers = [str(h).strip() if h else f"col_{i}" for i, h in enumerate(rows[0])]
    result = []
    for row in rows[1:]:
        if all(v is None for v in row):
            continue
        result.append({headers[i]: v for i, v in enumerate(row)})
    wb.close()
    return result

def _read_csv(path: str) -> list[dict]:
    with open(path, newline="", encoding="utf-8-sig") as f:
        return list(csv.DictReader(f))


# ── Main public API ───────────────────────────────────────────────────────────
def load_records(doc_type: str,
                 data_path: str | None = None,
                 validate: bool = True,
                 strict: bool = False) -> tuple[list[dict], list[str]]:
    """
    Load, coerce and optionally validate all records for a document type.

    Returns:
        (records, all_errors)
        records    — list of clean row dicts
        all_errors — list of validation error strings (empty = all clean)
    """
    doc_schema = get_doc_schema(doc_type)
    fields     = doc_schema["fields"]

    if data_path is None:
        _DATA_FILES = {
            "bank_statement":    "bank_statements.xlsx",
            "insurance_policy":  "insurance_policies.xlsx",
            "telecom_bill":      "telecom_bills.xlsx",
            "payroll_statement": "payroll_statements.xlsx",
        }
        fname     = _DATA_FILES.get(doc_type, f"{doc_type}.xlsx")
        data_path = str(Path(__file__).parent / "data" / fname)

    path = str(data_path)
    raw_rows = _read_csv(path) if path.endswith(".csv") else \
               _read_excel_sheet(path, doc_schema.get("data_sheet"))

    all_errors: list[str] = []
    records:    list[dict] = []

    for i, raw in enumerate(raw_rows, start=2):  # row 2 = first data row
        row = {}

        # Coerce types
        for field_name, field_def in fields.items():
            if field_def.get("type") == "list":
                continue
            raw_val = raw.get(field_name)
            coerced = _coerce(raw_val, field_def)

            # Apply masking
            if field_def.get("mask") and coerced:
                coerced = mask_field(str(coerced), field_def["mask"])

            row[field_name] = coerced

        # Carry over any extra columns not in schema (e.g. output_filename)
        for k, v in raw.items():
            if k not in row:
                row[k] = v if v is not None else ""

        # Validate
        if validate:
            errs = validate_row(row, fields, row_num=i)
            all_errors.extend(errs)
            if errs and strict:
                raise ValidationError(errs)

        records.append(row)

    return records, all_errors


def load_record(doc_type: str, row_index: int,
                data_path: str | None = None) -> dict:
    """Load a single record by zero-based index."""
    records, errors = load_records(doc_type, data_path, validate=True, strict=False)
    if errors:
        log.warning("Validation warnings for row %d: %s", row_index, errors)
    if row_index >= len(records):
        raise IndexError(f"Row {row_index} not found (file has {len(records)} rows).")
    return records[row_index]


def stream_records(doc_type: str,
                   data_path: str | None = None,
                   chunk_size: int = 500) -> Iterator[list[dict]]:
    """
    Generator that yields chunks of records.
    Memory-efficient for large batch runs (100K+ records).
    """
    records, _ = load_records(doc_type, data_path, validate=False)
    for i in range(0, len(records), chunk_size):
        yield records[i: i + chunk_size]
