"""
create_sample_data.py  —  Run once to generate sample Excel files in data/
"""
import os
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

os.makedirs("data", exist_ok=True)

H_FONT  = Font(bold=True, color="FFFFFF", size=10)
H_FILL  = PatternFill("solid", fgColor="0D1117")
H_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)

def style_ws(ws, headers):
    ws.row_dimensions[1].height = 26
    for i, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=i, value=h)
        c.font = H_FONT; c.fill = H_FILL; c.alignment = H_ALIGN
        ws.column_dimensions[c.column_letter].width = max(16, len(h)+3)

# ── 1. Bank Statements ────────────────────────────────────────────────────────
wb = openpyxl.Workbook()
ws = wb.active; ws.title = "BankStatements"
cols = ["account_number","account_holder","account_type","statement_date",
        "period_from","period_to","opening_balance","closing_balance",
        "total_credits","total_debits","minimum_balance","interest_rate",
        "branch_name","ifsc_code","output_filename"]
style_ws(ws, cols)
ws.append(["ACC****4821","James Wilson","Savings","26 Mar 2026",
           "01 Mar 2026","31 Mar 2026",12500.00,8340.50,
           3200.00,7359.50,1000.00,3.5,
           "Downtown Branch","HDFC0001234","stmt_wilson_mar26.pdf"])
ws.append(["ACC****9033","Priya Sharma","Checking","26 Mar 2026",
           "01 Mar 2026","31 Mar 2026",500.00,-245.75,
           1500.00,2245.75,500.00,0,
           "Uptown Branch","HDFC0005678","stmt_sharma_mar26.pdf"])
ws.append(["ACC****7712","Carlos Mendez","Credit","26 Mar 2026",
           "01 Mar 2026","31 Mar 2026",0,3420.00,
           0,3420.00,0,18.9,
           "West Branch","HDFC0009012","stmt_mendez_mar26.pdf"])

# Transactions sheet
ws2 = wb.create_sheet("Transactions")
cols2 = ["account_number","date","description","amount","balance","type"]
style_ws(ws2, cols2)
for row in [
    ["ACC****4821","01 Mar 2026","Opening Balance","0","12500.00",""],
    ["ACC****4821","05 Mar 2026","Salary Credit","5000.00","17500.00","CR"],
    ["ACC****4821","08 Mar 2026","Rent Payment","-3500.00","14000.00","DR"],
    ["ACC****4821","12 Mar 2026","Grocery Store","-320.50","13679.50","DR"],
    ["ACC****4821","15 Mar 2026","Netflix Subscription","-15.99","13663.51","DR"],
    ["ACC****4821","20 Mar 2026","Freelance Payment","2200.00","15863.51","CR"],
    ["ACC****4821","25 Mar 2026","Utility Bill","-523.01","15340.50","DR"],
    ["ACC****4821","28 Mar 2026","ATM Withdrawal","-1000.00","14340.50","DR"],
    ["ACC****4821","31 Mar 2026","Service Charge","-6000.00","8340.50","DR"],

    ["ACC****9033","01 Mar 2026","Opening Balance","0","500.00",""],
    ["ACC****9033","03 Mar 2026","ATM Withdrawal","-200.00","300.00","DR"],
    ["ACC****9033","10 Mar 2026","Grocery","-145.75","154.25","DR"],
    ["ACC****9033","15 Mar 2026","PayPal Transfer","1500.00","1654.25","CR"],
    ["ACC****9033","22 Mar 2026","Car Payment","-900.00","754.25","DR"],
    ["ACC****9033","28 Mar 2026","Online Shopping","-1000.00","-245.75","DR"],
]:
    ws2.append(row)
wb.save("data/bank_statements.xlsx")
print("✅  data/bank_statements.xlsx")


# ── 2. Insurance Policies ─────────────────────────────────────────────────────
wb = openpyxl.Workbook()
ws = wb.active; ws.title = "InsurancePolicies"
cols = ["policy_number","policy_type","insured_name","insured_dob","insured_address",
        "effective_date","expiry_date","premium_amount","premium_frequency",
        "sum_assured","deductible","agent_name","agent_code","output_filename"]
style_ws(ws, cols)
ws.append(["POL-2024-00182","Health","James Wilson","15 Jun 1985",
           "42 Oak Street, Chicago IL 60601",
           "01 Apr 2025","31 Mar 2026",285.00,"Monthly",
           500000.00,1500.00,"Sarah Green","AG-3821","policy_wilson.pdf"])
ws.append(["POL-2024-00291","Auto","Priya Sharma","22 Nov 1990",
           "88 Maple Ave, Austin TX 78701",
           "01 Jan 2026","31 Dec 2026",148.50,"Monthly",
           75000.00,500.00,"Mike Johnson","AG-4411","policy_sharma.pdf"])
ws.append(["POL-2023-00847","Life","Carlos Mendez","08 Mar 1978",
           "15 Pine Road, Miami FL 33101",
           "01 Jul 2023","30 Jun 2028",512.00,"Quarterly",
           2000000.00,0,"Linda Ross","AG-2290","policy_mendez.pdf"])

ws2 = wb.create_sheet("Coverages")
cols2 = ["policy_number","coverage_name","coverage_limit","notes"]
style_ws(ws2, cols2)
for row in [
    ["POL-2024-00182","Hospitalization","$500,000","Room rent capped at $500/day"],
    ["POL-2024-00182","Critical Illness","$100,000","36 listed conditions"],
    ["POL-2024-00182","Outpatient","$5,000","Per policy year"],
    ["POL-2024-00291","Own Damage","$75,000","IDV basis"],
    ["POL-2024-00291","Third Party Liability","Unlimited","As per Motor Act"],
    ["POL-2024-00291","Personal Accident","$25,000","Driver + owner"],
    ["POL-2023-00847","Death Benefit","$2,000,000","Payable to nominee"],
    ["POL-2023-00847","Terminal Illness","$2,000,000","Accelerated benefit"],
]:
    ws2.append(row)
wb.save("data/insurance_policies.xlsx")
print("✅  data/insurance_policies.xlsx")


# ── 3. Telecom Bills ──────────────────────────────────────────────────────────
wb = openpyxl.Workbook()
ws = wb.active; ws.title = "TelecomBills"
cols = ["account_number","customer_name","customer_address","bill_number",
        "bill_date","due_date","billing_period","plan_name",
        "data_used_gb","data_limit_gb","calls_minutes","sms_count",
        "previous_balance","current_charges","taxes_fees","total_due",
        "days_overdue","autopay_enabled","output_filename"]
style_ws(ws, cols)
ws.append(["TCM-88821","James Wilson","42 Oak Street, Chicago IL 60601",
           "BILL-2026-03-88821","26 Mar 2026","10 Apr 2026","01–31 Mar 2026",
           "Unlimited Plus",18.4,20,420,38,
           0.00,89.99,8.10,98.09,0,"true","bill_wilson_mar26.pdf"])
ws.append(["TCM-44309","Priya Sharma","88 Maple Ave, Austin TX 78701",
           "BILL-2026-03-44309","26 Mar 2026","10 Apr 2026","01–31 Mar 2026",
           "Basic 5GB",5.1,5,210,12,
           45.00,39.99,3.60,88.59,15,"false","bill_sharma_mar26.pdf"])
ws.append(["TCM-61177","Carlos Mendez","15 Pine Road, Miami FL 33101",
           "BILL-2026-03-61177","26 Mar 2026","25 Feb 2026","01–31 Mar 2026",
           "Family Share 100GB",87.3,100,1840,203,
           180.00,159.99,14.40,354.39,45,"false","bill_mendez_mar26.pdf"])

ws2 = wb.create_sheet("Charges")
cols2 = ["bill_number","charge_description","amount"]
style_ws(ws2, cols2)
for row in [
    ["BILL-2026-03-88821","Monthly Plan — Unlimited Plus","$79.99"],
    ["BILL-2026-03-88821","International Roaming (3 days)","$10.00"],
    ["BILL-2026-03-88821","State Tax (8.25%)","$7.44"],
    ["BILL-2026-03-88821","Federal USF Charge","$0.66"],
    ["BILL-2026-03-44309","Monthly Plan — Basic 5GB","$29.99"],
    ["BILL-2026-03-44309","Data Overage (0.1 GB)","$10.00"],
    ["BILL-2026-03-44309","State Tax","$3.60"],
    ["BILL-2026-03-44309","Previous Balance","$45.00"],
    ["BILL-2026-03-61177","Family Share 100GB Plan","$139.99"],
    ["BILL-2026-03-61177","Additional Line x1","$20.00"],
    ["BILL-2026-03-61177","State Tax","$14.40"],
    ["BILL-2026-03-61177","Previous Balance Outstanding","$180.00"],
]:
    ws2.append(row)
wb.save("data/telecom_bills.xlsx")
print("✅  data/telecom_bills.xlsx")


# ── 4. Payroll Statements ─────────────────────────────────────────────────────
wb = openpyxl.Workbook()
ws = wb.active; ws.title = "PayrollStatements"
cols = ["employee_id","employee_name","designation","department",
        "pay_period","pay_date","bank_account","pan_number",
        "basic_salary","hra","special_allowance","gross_earnings",
        "pf_deduction","tax_deduction","total_deductions","net_pay",
        "days_worked","leaves_taken","total_working_days","lwp_days",
        "increment_applied","output_filename"]
style_ws(ws, cols)
ws.append(["EMP-1041","James Wilson","Senior Software Engineer","Engineering",
           "March 2026","31 Mar 2026","ACC****4821","ABCDE****F",
           8000.00,3200.00,1800.00,13000.00,
           960.00,1560.00,2520.00,10480.00,
           22,1,22,0,"false","payslip_wilson_mar26.pdf"])
ws.append(["EMP-0892","Priya Sharma","Product Manager","Product",
           "March 2026","31 Mar 2026","ACC****9033","PQRST****A",
           9500.00,3800.00,2200.00,15500.00,
           1140.00,2170.00,3310.00,12190.00,
           20,2,22,0,"true","payslip_sharma_mar26.pdf"])
ws.append(["EMP-1187","Carlos Mendez","UX Designer","Design",
           "March 2026","31 Mar 2026","ACC****7712","WXYZL****B",
           6000.00,2400.00,1200.00,9600.00,
           720.00,720.00,1440.00,8160.00,
           18,2,22,2,"false","payslip_mendez_mar26.pdf"])

ws2 = wb.create_sheet("EarningsDeductions")
cols2 = ["employee_id","item_type","description","amount"]
style_ws(ws2, cols2)
for row in [
    ["EMP-1041","earning","Basic Salary","$8,000.00"],
    ["EMP-1041","earning","House Rent Allowance","$3,200.00"],
    ["EMP-1041","earning","Special Allowance","$1,800.00"],
    ["EMP-1041","deduction","Provident Fund (12%)","$960.00"],
    ["EMP-1041","deduction","Income Tax (TDS)","$1,560.00"],
    ["EMP-0892","earning","Basic Salary","$9,500.00"],
    ["EMP-0892","earning","House Rent Allowance","$3,800.00"],
    ["EMP-0892","earning","Special Allowance","$2,200.00"],
    ["EMP-0892","deduction","Provident Fund (12%)","$1,140.00"],
    ["EMP-0892","deduction","Income Tax (TDS)","$2,170.00"],
    ["EMP-1187","earning","Basic Salary","$6,000.00"],
    ["EMP-1187","earning","HRA","$2,400.00"],
    ["EMP-1187","earning","Special Allowance","$1,200.00"],
    ["EMP-1187","deduction","Provident Fund (12%)","$720.00"],
    ["EMP-1187","deduction","Income Tax (TDS)","$720.00"],
    ["EMP-1187","deduction","LWP Deduction (2 days)","$545.45"],
]:
    ws2.append(row)
wb.save("data/payroll_statements.xlsx")
print("✅  data/payroll_statements.xlsx")
print("\n✅  All data files created in data/")
